import pandas as pd
import requests
import wget
import os
import glob
import openpyxl
import numpy as np
from datetime import datetime
from datetime import timedelta
from os import scandir, getcwd

def descargaDia():
    
    xlsxFiles = glob.glob('diario/*.xlsx')
    xlsxFilesConsolidado = glob.glob('consolidado/*.xlsx')

    for g in xlsxFiles:
        try:
            os.remove(g)
        except OSError as e:
            print(f"Error:{ e.strerror}")
            
    for h in xlsxFilesConsolidado:
        try:
            os.remove(h)
        except OSError as e:
            print(f"Error:{ e.strerror}")
    
    now = datetime.now()
    currentDate = now.date()
    
    year = (currentDate).strftime("%Y")
    month = (currentDate).strftime("%m")
    urlBase = "https://www.odepa.gob.cl/wp-content/uploads/" + year + "/" + month + "/Boletin_Diario_de_Frutas_y_Hortalizas_" + str(currentDate).replace("-","") + ".xlsx"
    
    print(urlBase)
    
    try:
        currentFile = requests.get(urlBase)
        open('diario/' + str(currentDate).replace("-","") + '.xlsx', 'wb').write(currentFile.content)
        print("Archivo diario descargado correctamente.")
    except:
        print("¡Archivo no descargado!")

def Ciclo():
    descargaDia()
    Archivos = lsExcel()
    if(len(Archivos) > 0):
        actualizarDiario(Archivos)
        consolidadoHortaliza()
        consolidadoFruta()
        deDiaria()
    else:
        print("No hay datos que actualizar")
    print("Ciclo completo")
    # Ciclo()

def actualizarDiario(Archivos):
    wb = openpyxl.load_workbook("Diccionario.xlsx")
    hojas_for_dict = wb.sheetnames
    hojas_for_dict
    
    Mercado = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[0])
    Mercado.to_dict(orient = "list")["Mercado  "]
    for i in range(len(Mercado.to_dict(orient = "list")["Mercado  "])):
        print('"' + Mercado.to_dict(orient = "list")["Mercado  "][i] + '"')
        
    Mercado_Dict = {
        'Lo Valledor':["Mercado Mayorista Lo Valledor de Santiago",13],
    'Vega Central Mapocho':["Vega Central Mapocho de Santiago",13],
    'Macroferia Talca':["Macroferia Regional de Talca",7],
    'Femacal':["Femacal de La Calera",5],
    'La Palmera':["Terminal La Palmera de La Serena",4],
    'Solcoagro':["Comercializadora del Agro de Limarí",4],
    'Vega Monumental':["Vega Monumental Concepción",8],
    'Lagunita Pto.Montt':["Feria Lagunitas de Puerto Montt",10],
    'Vega Modelo Temuco':["Vega Modelo de Temuco",9],
    'Agrochillan':["Terminal Hortofrutícola Agro Chillán",16],
    'Agronor':["Agrícola del Norte S.A. de Arica",15],
    'Mapocho Vta.dir':  ["Mapocho Venta Directa de Santiago",13]
    }
    
    Region_Dict = {
        13:"Metropolitana",
        15:"Arica y Parinacota",
        4:"Coquimbo",
        5:"Coquimbo",
        10:"Los Lagos",
        7:"Maule",
        16:"Ñuble",
        9:"La Araucanía",
        8:"Bíobío"
    }

    Mes = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[1])
    
    Mes_Dict = {}
    for i in range(len(Mes)):
        Mes_Dict[i] = Mes["Mes"][i]
    Mes_Dict

    Especie  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[2])
    Especie_Dict = {}
    for i in range(len(Especie)):
        Especie_Dict[Especie["Especie"][i]] = Especie["Clasificación"][i]
    
    Detalle  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[3])
    Detalle_Dict = {}
    for i in range(len(Detalle)):
        Detalle_Dict[Detalle["Detalle"][i]] = Detalle["Kg"][i]
    Detalle_Dict['$/bandeja 18 kilos empedrada'] = 18
    Detalle_Dict['$/caja 18 kilos importada'] = 18
    Detalle_Dict['$/malla 22 kilos'] = 22
    Detalle_Dict['$/paquete 2 kilos'] = 2
    Detalle_Dict['$/caja 5 kilos'] = 5
    Detalle_Dict['$/atado'] = 1   #Preguntar
    Detalle_Dict['$/caja 8 kilos'] = 8
    Detalle_Dict['$/malla 100 unidades'] = 10 #Preguntar
    Detalle_Dict['$/media docena de atados'] = 6 #PReguntar
    Detalle_Dict['$/bins (500 kilos)'] = 500
    Detalle_Dict['$/envase 1 kilo'] = 1
    Detalle_Dict['$/cien'] = 10 #Preguntar
    Detalle_Dict['$/docena'] = 1.2 #Preguntar
    Detalle_Dict['$/caja 14 kilos'] = 14 #Preguntar
    Detalle_Dict['$/cien en rama (volumen en unidades)'] = 1 #Preguntar
    Detalle_Dict['$/caja 20 kilos empedrada'] = 20
    Detalle_Dict['$/caja 17 kilos empedrada'] = 17
    
    Detalle  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[3])
    Frutas = []
    Hortalizas = []
    for i in Archivos:
        print('diario/' + str(i))
        wb = openpyxl.load_workbook('diario/' +str(i))
        hojas = wb.sheetnames
        hojas
        
        dict_auxiliar = {}
        for hoja in hojas:
            if("Frutas" in hoja):
                #Frutas.append(hoja)
                datos = pd.read_excel('diario/' + str(i), sheet_name=hoja, skiprows=8, skipfooter=1)
                mercado_list = hoja.split("_")[1]
                mercado = Mercado_Dict[mercado_list][0]
                region = Region_Dict[Mercado_Dict[mercado_list][1]]
                cod_reg = Mercado_Dict[mercado_list][1]
                fecha = SalidaFecha(i)
                tipo = "Fruta"
                #print(mercado,region,fecha, cod_reg)
                #print(len(datos))
                for filas in range(len(datos)):
                    categoria = Especie_Dict[datos["Producto "][filas]]
                    producto = datos["Producto "][filas]
                    variedad = datos['Variedad '][filas]
                    calidad = datos['Calidad '][filas]
                    volumen = datos['Volumen '][filas]
                    precio_minimo = datos['Precio\nmínimo'][filas]
                    precio_maximo = datos['Precio\nmáximo'][filas]
                    precio_promedio = datos['Precio\npromedio'][filas]
                    u_comercializacion = datos['Unidad de\ncomercialización '][filas]
                    origen = datos['Origen '][filas]
                    try:
                        kgUnidad = Detalle_Dict[u_comercializacion]
                    except:
                        kgUnidad = 1
                    
                    
                    precio = int(round(precio_promedio / kgUnidad,0))               
                    Frutas.append(diccionario_auxiliar(mercado,region,fecha,cod_reg,tipo,categoria,producto,variedad,calidad,volumen,precio_minimo,precio_maximo,precio_promedio,u_comercializacion,origen, precio,kgUnidad))
                    #Frutas.append(
                #break
        
        for hoja in hojas:
            if("Hortalizas" in hoja):
                #Frutas.append(hoja)
                datos = pd.read_excel('diario/' + str(i), sheet_name=hoja, skiprows=8, skipfooter=1)
                mercado_list = hoja.split("_")[1]
                mercado = Mercado_Dict[mercado_list][0]
                region = Region_Dict[Mercado_Dict[mercado_list][1]]
                cod_reg = Mercado_Dict[mercado_list][1]
                fecha = SalidaFecha(i)
                tipo = "Fruta"
                #print(mercado,region,fecha, cod_reg)
                #print(len(datos))
                for filas in range(len(datos)):
                    categoria = ""
                    producto = datos["Producto "][filas]
                    variedad = datos['Variedad '][filas]
                    calidad = datos['Calidad '][filas]
                    volumen = datos['Volumen '][filas]
                    precio_minimo = datos['Precio\nmínimo'][filas]
                    precio_maximo = datos['Precio\nmáximo'][filas]
                    precio_promedio = datos['Precio\npromedio'][filas]
                    try:
                        u_comercializacion = datos['Unidad de\ncomercialización '][filas]
                    except:
                        u_comercializacion = datos['Unidad de\ncomercialización'][filas]
                    origen = datos['Origen '][filas]
                    try:
                        kgUnidad = Detalle_Dict[u_comercializacion]
                    except:
                        kgUnidad = 1
                    #kgUnidad = Detalle_Dict[u_comercializacion]
                    precio = int(round(precio_promedio / kgUnidad,0))               
                    Hortalizas.append(diccionario_auxiliar(mercado,region,fecha,cod_reg,tipo,categoria,producto,variedad,calidad,volumen,precio_minimo,precio_maximo,precio_promedio,u_comercializacion,origen, precio,kgUnidad))
    datosFruta = pd.DataFrame(Frutas)
    datosHortaliza = pd.DataFrame(Hortalizas)
    
    datosHortaliza["Kg o Unidades"] = datosHortaliza["Kg / unidad"]
    del datosHortaliza["Tipo"]
    del datosHortaliza["Categoría"]
    del datosHortaliza["Kg / unidad"]

    fruta_salida = pd.concat([ref_frutas1(),datosFruta])
    hortaliza_salida = pd.concat([ref_hortalizas1(),datosHortaliza])
    fruta_salida.fillna(0)
    hortaliza_salida.fillna(0)
    #fruta_salida = datosFruta
    #hortaliza_salida = datosHortaliza
    hortaliza_salida["Clasificación"] = "Hortaliza"

    # Archivos consolidados HECTOR
    fruta_salida.to_excel("Consolidado/FrutaConsolidado.xlsx", index=False)
    hortaliza_salida.to_excel("Consolidado/HortalizaConsolidado.xlsx", index=False)
    return

def lsExcel():
    salida = []
    # for i in ls("PrecioFrutaHortalizas"):
    for i in ls("diario"):
        if("xlsx" in i and "20" in i and ".tmp" not in i and i[0] == "2"):
            salida.append(i)
    return salida

def ls(ruta = getcwd()):
    return [arch.name for arch in scandir(ruta) if arch.is_file()]

def SalidaFecha(nombre):

    fecha_str = nombre[6:8] + "-" + nombre[4:6] + "-" + nombre[0:4]
    return datetime.strptime(fecha_str,"%d-%m-%Y")

def diccionario_auxiliar(Mercado, Region, Fecha, Codreg, Tipo, Categoria, Producto,
       Variedad, Calidad, Volumen, Precio_minimo, Precio_maximo,
       Precio_promedio_ponderado, Unidad_de_comercializacion, Origen,
       Precio_Kg, Kg_unidad):
    return {'Mercado' : Mercado, 
        'Región' : Region, 
        'Fecha' : Fecha, 
        'Codreg' : Codreg, 
        'Tipo' : Tipo, 
        'Categoría' : Categoria, 
        'Producto' : Producto,
        'Variedad' : Variedad, 
        'Calidad' : Calidad, 
        'Volumen' : Volumen, 
        'Precio mínimo' : Precio_minimo, 
        'Precio máximo' : Precio_maximo,
       'Precio promedio ponderado' : Precio_promedio_ponderado, 
        'Unidad de comercialización' : Unidad_de_comercializacion, 
        'Origen':Origen,
       'Precio $/Kg' : Precio_Kg, 
        'Kg / unidad' : Kg_unidad}

def ref_frutas1():
    dataReferenciaFruta = pd.read_excel("FrutaConsolidado1.xlsx")
    return dataReferenciaFruta

def ref_hortalizas1():
    dataReferenciaHortaliza = pd.read_excel("HortalizaConsolidado1.xlsx")
    return dataReferenciaHortaliza

dfC = pd.read_excel("Consolidado/FrutaConsolidado.xlsx")
dfH = pd.read_excel("Consolidado/HortalizaConsolidado.xlsx")

referenciaProd = pd.read_excel("referenciaProducto.xlsx") 
referenciaCate = pd.read_excel("referenciaCategoría.xlsx") 

def consolidadoFruta():
    print("Creando consolidado Frutas")   
    
    datos = []

    for i, index in dfC.iterrows():
    
        prodReferencia = referenciaProd
        cateReferencia = referenciaCate
        
        _prod = dfC["Categoría"][i]
        _cate = dfC["Producto"][i]

        
        if (_prod == "Oleaginosos"):
            _prod = "Frutos oleaginosos"

        elif(_prod == "Breva"):
            _prod = "Higo"

        elif(_prod == "Haba"):
            _prod = "Habas"

        elif(_prod == "Sandia"):
            _prod = "Sandía"

        else:
            pass

        
        prodReferencia = prodReferencia[prodReferencia["nombre"] == str(_prod)]

        try:
            idD = prodReferencia["id"].to_list()
            
            # print(_prod)
            # print("Producto ID: " + str(idD[0]))
        except:
            idD = ""
            
        cateReferencia = cateReferencia[cateReferencia["nombre"] == str(_cate)]

        try:
            idP = cateReferencia["id"].to_list()
            
            mer = dfC["Mercado"][i]
            reg = dfC["Región"][i]
            fec = dfC["Fecha"][i]
            codR = dfC["Codreg"][i]
            tipo = dfC["Tipo"][i]
            cate = dfC["Categoría"][i]
            prod = dfC["Producto"][i]
            var = dfC["Variedad"][i]
            cal = dfC["Calidad"][i]
            vol = dfC["Volumen"][i]
            pmin = dfC["Precio mínimo"][i]
            pm = dfC["Precio máximo"][i]
            ppp = dfC["Precio promedio ponderado"][i]
            uc = dfC["Unidad de comercialización"][i]
            ori = dfC["Origen"][i]
            pkg = dfC["Precio $/Kg"][i]
            kgu = dfC["Kg / unidad"][i]

            merId = mercadoID(dfC["Mercado"][i])

            diccionario = registros(merId, mer, reg, fec, codR, tipo, idD[0], cate, idP[0], prod, var, cal, vol, pmin, pm, ppp, uc, ori, pkg, kgu)
            datos.append(diccionario.copy())
            # print(_cate)
            # print("Categoría ID: " + str(idP[0]))
        except:
            idP = ""

            mer = dfC["Mercado"][i]
            reg = dfC["Región"][i]
            fec = dfC["Fecha"][i]
            codR = dfC["Codreg"][i]
            tipo = dfC["Tipo"][i]
            cate = dfC["Categoría"][i]
            prod = dfC["Producto"][i]
            var = dfC["Variedad"][i]
            cal = dfC["Calidad"][i]
            vol = dfC["Volumen"][i]
            pmin = dfC["Precio mínimo"][i]
            pm = dfC["Precio máximo"][i]
            ppp = dfC["Precio promedio ponderado"][i]
            uc = dfC["Unidad de comercialización"][i]
            ori = dfC["Origen"][i]
            pkg = dfC["Precio $/Kg"][i]
            kgu = dfC["Kg / unidad"][i]

            merId = mercadoID(dfC["Mercado"][i])

            diccionario = registros(merId, mer, reg, fec, codR, tipo, idD, cate, idP, prod, var, cal, vol, pmin, pm, ppp, uc, ori, pkg, kgu)
            datos.append(diccionario.copy())

        # print(idD)
        # print(idP)

    data = pd.DataFrame(datos)
    data.to_excel("Consolidado/FrutaConsolidado.xlsx", index=False)
    print("Consolidado Frutas")

def consolidadoHortaliza():

    print("Creando consolidado Hortalizas")

    datos = []

    for i, index in dfH.iterrows():
    
        cateReferencia = referenciaCate
        
        _cate = dfH["Producto"][i]

        
        if (_cate == "Oleaginosos"):
            _cate = "Frutos oleaginosos"

        elif(_cate == "Breva"):
            _cate = "Higo"

        elif(_cate == "Haba"):
            _cate = "Habas"

        elif(_cate == "Sandia"):
            _cate = "Sandía"

        
        cateReferencia = cateReferencia[cateReferencia["nombre"] == str(_cate)]

        try:
            idD = cateReferencia["id"].to_list()
            
            mer = dfH["Mercado"][i]
            reg = dfH["Región"][i]
            fec = dfH["Fecha"][i]
            codR = dfH["Codreg"][i]
            prod = dfH["Producto"][i]
            var = dfH["Variedad"][i]
            cal = dfH["Calidad"][i]
            vol = dfH["Volumen"][i]
            pmin = dfH["Precio mínimo"][i]
            pm = dfH["Precio máximo"][i]
            ppp = dfH["Precio promedio ponderado"][i]
            uc = dfH["Unidad de comercialización"][i]
            ori = dfH["Origen"][i]
            pkg = dfH["Precio $/Kg"][i]
            kgu = dfH["Kg o Unidades"][i]
            clasi = dfH["Clasificación"][i]

            merId = mercadoID(dfH["Mercado"][i])

            diccionario = registros2(merId, mer, reg, fec, codR, idD[0], prod, var, cal, vol, pmin, pm, ppp, uc, ori, pkg, kgu, clasi)
            datos.append(diccionario.copy())

            # print(_prod)
            # print("Categoría ID: " + str(idD[0]))
        except:
            idD = ""


            mer = dfH["Mercado"][i]
            reg = dfH["Región"][i]
            fec = dfH["Fecha"][i]
            codR = dfH["Codreg"][i]
            prod = dfH["Producto"][i]
            var = dfH["Variedad"][i]
            cal = dfH["Calidad"][i]
            vol = dfH["Volumen"][i]
            pmin = dfH["Precio mínimo"][i]
            pm = dfH["Precio máximo"][i]
            ppp = dfH["Precio promedio ponderado"][i]
            uc = dfH["Unidad de comercialización"][i]
            ori = dfH["Origen"][i]
            pkg = dfH["Precio $/Kg"][i]
            kgu = dfH["Kg o Unidades"][i]
            clasi = dfH["Clasificación"][i]

            merId = mercadoID(dfH["Mercado"][i])

            diccionario = registros2(merId, mer, reg, fec, codR, idD, prod, var, cal, vol, pmin, pm, ppp, uc, ori, pkg, kgu, clasi)
            datos.append(diccionario.copy())

        # print(idD)
        # print(idP)

    data = pd.DataFrame(datos)
    data.to_excel("Consolidado/HortalizaConsolidado.xlsx", index=False)
    print("Consolidado Hortalizas")

_mercadoID = {'Agrícola del Norte S.A. de Arica':'1', 
              'Comercializadora del Agro de Limarí':'2',
              'Femacal de La Calera':'3', 
              'Feria Lagunitas de Puerto Montt':'4',
              'Macroferia Regional de Talca':'5', 
              'Mercado Mayorista Lo Valledor de Santiago':'6',
              'Terminal Hortofrutícola Agro Chillán':'7', 
              'Terminal La Palmera de La Serena':'8',
              'Vega Central Mapocho de Santiago':'9', 
              'Vega Modelo de Temuco':'10',
              'Vega Monumental Concepción':'11', 
              'Mapocho Venta Directa de Santiago':'12'}

def mercadoID(mercado):
    
    value = 0
    value = _mercadoID[mercado]
        
    return value

def registros(meID, Mercado, Region, Fecha, Codreg, Tipo, cateID, Categoria, prodID, Producto, Variedad, Calidad, Volumen, PrecioMin, PrecioMax, ppp, UnidadComer, Origen, PrecioKg, KgUnidad):
    diccionario = {}
    diccionario["Mercado ID"] = meID
    diccionario["Mercado"] = Mercado
    diccionario["Región"] = Region
    diccionario["Fecha"] = Fecha
    diccionario["Codreg"] = Codreg
    diccionario["Tipo"] = Tipo
    diccionario["Producto ID"] = cateID
    diccionario["Producto"] = Categoria
    diccionario["Categoría ID"] = prodID
    diccionario["Categoría"] = Producto
    diccionario["Variedad"] = Variedad
    diccionario["Calidad"] = Calidad
    diccionario["Volumen"] = Volumen
    diccionario["Precio mínimo"] = PrecioMin
    diccionario["Precio máximo"] = PrecioMax
    diccionario["Precio promedio ponderado"] = ppp
    diccionario["Unidad de comercialización"] = UnidadComer
    diccionario["Origen"] = Origen
    diccionario["Precio $/Kg"] = PrecioKg
    diccionario["Kg / unidad"] = KgUnidad

    return diccionario

def registros2(meID, Mercado, Region, Fecha, Codreg, prodID, Producto, Variedad, Calidad, Volumen, PrecioMin, PrecioMax, ppp, UnidadComer, Origen, PrecioKg, KgUnidad, clasi):
    diccionario = {}
    diccionario["Mercado ID"] = meID
    diccionario["Mercado"] = Mercado
    diccionario["Región"] = Region
    diccionario["Fecha"] = Fecha
    diccionario["Codreg"] = Codreg
    diccionario["Categoría ID"] = prodID
    diccionario["Categoría"] = Producto
    diccionario["Variedad"] = Variedad
    diccionario["Calidad"] = Calidad
    diccionario["Volumen"] = Volumen
    diccionario["Precio mínimo"] = PrecioMin
    diccionario["Precio máximo"] = PrecioMax
    diccionario["Precio promedio ponderado"] = ppp
    diccionario["Unidad de comercialización"] = UnidadComer
    diccionario["Origen"] = Origen
    diccionario["Precio $/Kg"] = PrecioKg
    diccionario["Kg o Unidades"] = KgUnidad
    diccionario["Clasificación"] = clasi

    return diccionario

def deDiaria():
    fileDay = "consolidado/*.xlsx"
    fileDay = glob.glob(fileDay)

    salidaUpdateF = []
    salidaUpdateH = []

    archivosDay = np.array(fileDay)
    count = 0

    if(len(archivosDay) > 0):

        for h in archivosDay:
            
            _dfDay = pd.read_excel(h)

            file = "subcojuntos/*.xlsx"
            files = glob.glob(file)

            archivos = np.array(files)
            archivos

            for i in archivos:

                dfDay = _dfDay

                df = pd.read_excel(i)

                std = df["Precio $/Kg"].std()
                _min = df["Precio $/Kg"].mean() - 3 * std
                _max = df["Precio $/Kg"].mean() + 3 * std

                dfAux = df[df["Precio $/Kg"] < df["Precio $/Kg"].mean() + 3 * std]
                dfAux = dfAux[dfAux["Precio $/Kg"] > dfAux["Precio $/Kg"].mean() - 3 * std]

                # print(std)


                try:
                    mercado = dfAux["Mercado"][0]
                    categoria = dfAux["Categoría"][0]

                    # print(mercado)
                    # print(categoria)

                    dfMer = dfDay[dfDay["Mercado"] == str(mercado)]
                    dfCate = dfMer[dfMer["Categoría"] == str(categoria)]      
                    stdDay = dfCate["Precio $/Kg"].std()

                    if(stdDay > 0):
                        # print(i)
                        # print("MIN: " + str(_min))
                        # print("MÁX: " + str(_max))
                        # print("")
                        # print("Nueva DE: " + str(stdDay))

                        if(stdDay >= _min) and (stdDay <= _max):
                            # print("SE AGREGA")
                            if(count == 0):
                                # print("Agregado frutas")
                                salidaUpdateF.append(dfCate.copy())
                            else:
                                # print("Agregado hortalzias")
                                salidaUpdateH.append(dfCate.copy())
                        else:
                            # print("NO SE AGREGA")
                            pass

                        # print("")
                    else:
                        pass

                except:
                    pass
            count += 1

        dataUpdateF = pd.concat(salidaUpdateF)
        dataUpdateH = pd.concat(salidaUpdateH)
        
        # dataUpdateF.to_excel("FRUTA.xlsx", index=False)
        # dataUpdateH.to_excel("HORTALIZA.xlsx", index=False)
        
        dfFrutas = pd.read_excel("FrutaConsolidado.xlsx")
        dfHortalizas = pd.read_excel("HortalizaConsolidado.xlsx")
        
        finalF = pd.concat([dfFrutas, dataUpdateF])
        finalH = pd.concat([dfHortalizas, dataUpdateH])
        
        finalF.to_excel("FrutaConsolidado.xlsx", index=False)
        finalH.to_excel("HortalizaConsolidado.xlsx", index=False)
        
        
    else:

        print("No se ha actualzado el archivo: no hay registros diarios.")
    
    print("Desvación esándar diaria aplicada correctamente.")
    # Validación si hay productos nuevos
            
if __name__ == '__main__':
    print('El proceso ha comenzado.')
    Ciclo()
    print('El proceso ha finalizado.')